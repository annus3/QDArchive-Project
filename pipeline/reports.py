"""
pipeline/reports.py

* ``export_xlsx``       — the six-column classification table (Step 4c) via openpyxl,
                          styled (frozen header, auto-filter, banded rows) with an
                          extra per-repository "Summary" sheet.
* ``generate_pdf``      — a **vector** PDF (Step 4d)
                          via reportlab: cover page, executive summary, stat cards,
                          project-type tables, primary-class histograms (full class
                          names as bins, counts on bars), top-20 ranked tables, and
                          comments.
* ``print_statistics``  — per-repository project-type counts + dominant class
                          (Step 4b) and a data-challenges summary (Step 5).

Report/export dependencies (``openpyxl``, ``reportlab``) are imported inside
the functions so importing ``pipeline`` never requires them.
"""

from __future__ import annotations

import logging
import os

from . import config
from .classifier import load_name_map, display_class
from .database import Database

logger = logging.getLogger(__name__)

_XLSX_COLUMNS = [
    "repository_id", "project_type", "project_title",
    "primary_class", "secondary_class", "no_project_files",
]

# Shared colour palette (used by both XLSX and PDF for a consistent look).
_NAVY = "1B3A5C"
_BLUE = "2E6DA4"
_GOLD = "E8A020"
_LGRAY = "F5F7FA"
_MGRAY = "CCCCCC"
_DGRAY = "555555"
_BAR_PALETTE = [
    "1B3A5C", "2E6DA4", "4A9FD4", "6AB8E8", "E8A020",
    "C0392B", "27AE60", "8E44AD", "E67E22", "16A085",
    "2980B9", "D35400", "7F8C8D", "2C3E50", "F39C12",
]


def _repos(repos: list[str] | None) -> list[str]:
    return repos or [k for k, v in config.REPOSITORIES.items() if v.get("enabled")]


def _repo_view(db: Database, repo_key: str, name_map: dict) -> dict:
    """Aggregate everything the reports need for one repository."""
    cfg = config.REPOSITORIES.get(repo_key, {})
    stats = db.classification_stats(repo_key, config.CLASSIFIABLE_TYPES)
    dist = db.class_distribution(repo_key, config.CLASSIFIABLE_TYPES)
    type_counts = stats["type_counts"]
    total = sum(v for k, v in type_counts.items() if k in config.PROJECT_TYPES)
    classified = sum(c for _, c in dist)
    return {
        "key": repo_key,
        "name": cfg.get("name", repo_key),
        "repository_id": cfg.get("repository_id"),
        "type_counts": type_counts,
        "total": total,
        "dist": dist,                       # [(code, count)] desc
        "n_classes": len(dist),
        "classified": classified,
        "dominant_code": stats["dominant_class"],
        "dominant_count": stats["dominant_count"],
    }


# ── Step 4c: XLSX table ─────────────────────────────────────────────────────
def export_xlsx(db: Database, path: str, repos: list[str] | None = None) -> str:
    """Write the classification table (QDA + QD projects) as a styled XLSX."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    name_map = load_name_map()
    repos = _repos(repos)
    os.makedirs(os.path.dirname(path), exist_ok=True)

    header_fill = PatternFill("solid", fgColor=_NAVY)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    band_fill = PatternFill("solid", fgColor=_LGRAY)
    thin = Side(style="thin", color=_MGRAY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    wb = Workbook()

    # -- Main sheet: the exact six required columns -------------------------
    ws = wb.active
    ws.title = "Classification"
    ws.append(_XLSX_COLUMNS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    row_idx = 1
    for repo_key in repos:
        for r in db.classification_table_rows(repo_key, config.CLASSIFIABLE_TYPES):
            row_idx += 1
            values = [
                r["repository_id"],
                r["project_type"],
                r["project_title"],
                display_class(r["primary_class"], name_map),
                display_class(r["secondary_class"], name_map),
                r["no_project_files"],
            ]
            ws.append(values)
            for col in range(1, len(values) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border
                cell.alignment = center if col in (1, 2, 6) else left
                if row_idx % 2 == 0:
                    cell.fill = band_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_XLSX_COLUMNS))}{row_idx}"
    widths = [14, 16, 60, 42, 42, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # -- Summary sheet: per-repo type counts + top classes ------------------
    sm = wb.create_sheet("Summary")
    sm.append(["Classification Summary"])
    sm["A1"].font = Font(bold=True, size=14, color=_NAVY)
    sm.append([])
    for repo_key in repos:
        v = _repo_view(db, repo_key, name_map)
        sm.append([f"{v['name']}  (repository_id={v['repository_id']})"])
        sm.cell(row=sm.max_row, column=1).font = Font(bold=True, color=_BLUE, size=12)
        sm.append(["Project type", "Count"])
        for c in sm[sm.max_row]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
        for t in config.PROJECT_TYPES:
            if t in v["type_counts"]:
                sm.append([t, v["type_counts"][t]])
        sm.append(["Dominant class",
                   display_class(v["dominant_code"], name_map) or "—"])
        sm.append(["Top primary classes", "Count"])
        for c in sm[sm.max_row]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
        for code, cnt in v["dist"][:10]:
            sm.append([name_map.get(code, code), cnt])
        sm.append([])
    sm.column_dimensions["A"].width = 52
    sm.column_dimensions["B"].width = 14

    wb.save(path)
    logger.info("Wrote XLSX → %s", path)
    return path


# ── Step 4d: professional vector PDF (reportlab) ────────────────────────────
def generate_pdf(db: Database, path: str, repos: list[str] | None = None,
                 top_n: int = 20) -> str:
    from reportlab.lib import colors
    from reportlab.lib.colors import HexColor
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        BaseDocTemplate, Frame, PageTemplate, NextPageTemplate, Paragraph,
        Spacer, Table, TableStyle, PageBreak, HRFlowable, Flowable, KeepTogether,
    )
    from reportlab.graphics.shapes import Drawing, Rect, String, Line
    from reportlab.pdfbase.pdfmetrics import stringWidth

    name_map = load_name_map()
    repos = _repos(repos)
    views = [_repo_view(db, rk, name_map) for rk in repos]
    os.makedirs(os.path.dirname(path), exist_ok=True)

    navy, blue, gold = HexColor("#" + _NAVY), HexColor("#" + _BLUE), HexColor("#" + _GOLD)
    lgray, mgray, dgray = HexColor("#" + _LGRAY), HexColor("#" + _MGRAY), HexColor("#" + _DGRAY)

    styles = {
        "h1": ParagraphStyle("h1", fontName="Helvetica-Bold", fontSize=15,
                             textColor=navy, spaceBefore=14, spaceAfter=6),
        "h2": ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=12,
                             textColor=blue, spaceBefore=10, spaceAfter=10),
        "body": ParagraphStyle("body", fontName="Helvetica", fontSize=10,
                               leading=15, textColor=dgray, alignment=TA_JUSTIFY,
                               spaceAfter=6),
        "cap": ParagraphStyle("cap", fontName="Helvetica-Oblique", fontSize=8,
                              textColor=HexColor("#888888"), alignment=TA_CENTER,
                              spaceAfter=6),
    }

    def table_style(extra=None):
        base = [
            ("BACKGROUND", (0, 0), (-1, 0), navy),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, lgray]),
            ("GRID", (0, 0), (-1, -1), 0.4, mgray),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
        return TableStyle(base + (extra or []))

    def wrap_text(text, max_w, font="Helvetica", size=6.5, max_lines=2):
        """Word-wrap to at most max_lines measured lines (spec 4d requires the
        FULL class name as the bin label, so wrapping replaces truncation)."""
        words = text.split()
        lines, cur = [], ""
        for word in words:
            trial = f"{cur} {word}".strip()
            if not cur or stringWidth(trial, font, size) <= max_w:
                cur = trial
            else:
                lines.append(cur)
                cur = word
        lines.append(cur)
        if len(lines) > max_lines:                # last-resort fallback only
            lines = lines[:max_lines]
            while lines[-1] and stringWidth(lines[-1] + "…", font, size) > max_w:
                lines[-1] = lines[-1][:-1].rstrip()
            lines[-1] += "…"
        return lines

    def hbar_chart(dist):
        # Untitled by design: the section heading in the story carries the
        # repository name / top-N context instead of an in-chart title.
        items = dist[:top_n]
        if not items:
            return None
        labels = [name_map.get(c, c) for c, _ in items]
        values = [n for _, n in items]
        n = len(items)
        bar_h, gap = 15, 5
        left_m, right_m, top_m, bot_m = 230, 55, 8, 22
        chart_w = 280
        W = left_m + chart_w + right_m
        H = top_m + n * (bar_h + gap) + bot_m
        d = Drawing(W, H)
        d.hAlign = "CENTER"
        maxv = max(values) or 1
        ink = HexColor("#333333")
        for i, (label, val) in enumerate(zip(labels, values)):
            y = bot_m + (n - 1 - i) * (bar_h + gap)
            blen = val / maxv * chart_w
            color = HexColor("#" + _BAR_PALETTE[i % len(_BAR_PALETTE)])
            d.add(Rect(left_m, y, blen, bar_h, fillColor=color,
                       strokeColor=colors.white, strokeWidth=0.4))
            lines = wrap_text(label, left_m - 14)
            if len(lines) == 1:
                d.add(String(left_m - 6, y + bar_h / 2 - 2.5, lines[0],
                             fontName="Helvetica", fontSize=6.5,
                             fillColor=ink, textAnchor="end"))
            else:
                d.add(String(left_m - 6, y + bar_h / 2 + 1.4, lines[0],
                             fontName="Helvetica", fontSize=6.5,
                             fillColor=ink, textAnchor="end"))
                d.add(String(left_m - 6, y + bar_h / 2 - 5.6, lines[1],
                             fontName="Helvetica", fontSize=6.5,
                             fillColor=ink, textAnchor="end"))
            d.add(String(left_m + blen + 4, y + bar_h / 2 - 3, f"{val:,}",
                         fontName="Helvetica-Bold", fontSize=7,
                         fillColor=ink, textAnchor="start"))
        d.add(Line(left_m, bot_m - 2, left_m + chart_w, bot_m - 2,
                   strokeColor=mgray, strokeWidth=0.5))
        return d

    def card_row(cards):
        """A row of KPI stat tiles (label under a large value)."""
        cells = [
            Paragraph(f'<font size="19"><b>{val}</b></font><br/>'
                      f'<font size="8" color="#666666">{lbl}</font>',
                      ParagraphStyle("s", alignment=TA_CENTER, leading=22))
            for lbl, val in cards
        ]
        box_w = (A4[0] - 3.6 * cm) / len(cards)
        t = Table([cells], colWidths=[box_w] * len(cards))
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), lgray),
            ("BOX", (0, 0), (-1, -1), 1, blue),
            ("LINEAFTER", (0, 0), (-2, -1), 0.5, mgray),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        return t

    def stat_cards(v):
        return card_row([
            ("Total projects", f"{v['total']:,}"),
            ("QDA_PROJECT", f"{v['type_counts'].get(config.PROJECT_TYPE_QDA, 0):,}"),
            ("QD_PROJECT", f"{v['type_counts'].get(config.PROJECT_TYPE_QD, 0):,}"),
            ("ISIC classes", f"{v['n_classes']:,}"),
        ])

    # Categorical palette for the four project types (fixed slot order,
    # CVD-validated ≥ 3:1 pair separation; low-contrast slots get direct labels).
    type_colors = {
        config.PROJECT_TYPE_QDA: HexColor("#2A78D6"),    # blue
        config.PROJECT_TYPE_QD: HexColor("#1BAF7A"),     # aqua
        config.PROJECT_TYPE_OTHER: HexColor("#EDA100"),  # yellow
        config.PROJECT_TYPE_NONE: HexColor("#4A3AA7"),   # violet
    }
    _dark_ink = HexColor("#333333")

    def type_composition_chart(views_):
        """100%-stacked horizontal bar per repository: project-type composition."""
        rows = [(v["name"], v["type_counts"], v["total"]) for v in views_ if v["total"]]
        if not rows:
            return None
        label_w, chart_w, right_m = 130, 300, 50
        bar_h, row_gap, top_pad, bot_pad = 18, 12, 4, 4
        n = len(rows)
        W = label_w + chart_w + right_m
        H = top_pad + 10 + 12 + n * bar_h + (n - 1) * row_gap + bot_pad
        d = Drawing(W, H)
        d.hAlign = "CENTER"
        # Legend (fixed type order) above the bars.
        lx, ly = label_w, H - top_pad - 7
        for t in config.PROJECT_TYPES:
            d.add(Rect(lx, ly - 1, 7, 7, fillColor=type_colors[t], strokeColor=None))
            d.add(String(lx + 10, ly, t, fontName="Helvetica", fontSize=7,
                         fillColor=_dark_ink))
            lx += 10 + stringWidth(t, "Helvetica", 7) + 14
        y = ly - 12 - bar_h
        for name, counts, total in rows:
            d.add(String(label_w - 8, y + bar_h / 2 - 3, name,
                         fontName="Helvetica-Bold", fontSize=8,
                         fillColor=_dark_ink, textAnchor="end"))
            x = label_w
            for t in config.PROJECT_TYPES:
                cnt = counts.get(t, 0)
                if not cnt:
                    continue
                seg = cnt / total * chart_w
                d.add(Rect(x, y, seg, bar_h, fillColor=type_colors[t],
                           strokeColor=colors.white, strokeWidth=1))
                # Selective direct labels: skip slivers (legend + table carry them).
                if seg >= 24:
                    pct = cnt / total * 100
                    lab = f"{pct:.0f}% ({cnt:,})" if seg >= 62 else f"{pct:.0f}%"
                    ink = (colors.white
                           if t in (config.PROJECT_TYPE_QDA, config.PROJECT_TYPE_NONE)
                           else _dark_ink)
                    d.add(String(x + seg / 2, y + bar_h / 2 - 2.5, lab,
                                 fontName="Helvetica-Bold", fontSize=7,
                                 fillColor=ink, textAnchor="middle"))
                x += seg
            d.add(String(label_w + chart_w + 6, y + bar_h / 2 - 3, f"{total:,}",
                         fontName="Helvetica-Bold", fontSize=7.5,
                         fillColor=_dark_ink, textAnchor="start"))
            y -= bar_h + row_gap
        return d

    def coverage_meters(rows):
        """Single-ratio meters: label · track+fill · 'done / total (pct)'."""
        label_w, track_w, right_m = 190, 190, 100
        row_h, row_gap, pad = 12, 12, 4
        n = len(rows)
        W = label_w + track_w + right_m
        H = 2 * pad + n * row_h + (n - 1) * row_gap
        d = Drawing(W, H)
        d.hAlign = "CENTER"
        y = H - pad - row_h
        for label, done, total in rows:
            frac = (done / total) if total else 0.0
            d.add(String(label_w - 8, y + row_h / 2 - 3, label, fontName="Helvetica",
                         fontSize=8, fillColor=_dark_ink, textAnchor="end"))
            d.add(Rect(label_w, y, track_w, row_h, fillColor=HexColor("#E4EAF2"),
                       strokeColor=None))
            if frac > 0:
                d.add(Rect(label_w, y, max(2, track_w * frac), row_h,
                           fillColor=HexColor("#2A78D6"), strokeColor=None))
            d.add(String(label_w + track_w + 8, y + row_h / 2 - 3,
                         f"{done:,} / {total:,}  ({frac * 100:.1f}%)",
                         fontName="Helvetica-Bold", fontSize=8,
                         fillColor=_dark_ink, textAnchor="start"))
            y -= row_h + row_gap
        return d

    # -- Page furniture ------------------------------------------------------
    def on_page(canvas, doc):
        w, h = A4
        canvas.saveState()
        canvas.setFillColor(navy)
        canvas.rect(0, h - 1.2 * cm, w, 1.2 * cm, fill=1, stroke=0)
        canvas.setFillColor(gold)
        canvas.rect(0, h - 1.34 * cm, w, 0.14 * cm, fill=1, stroke=0)
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 9)
        canvas.drawString(1.6 * cm, h - 0.82 * cm,
                          "QDArchive Classification Report")
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(w - 1.6 * cm, h - 0.82 * cm, "ISIC Rev. 5")
        canvas.setFillColor(navy)
        canvas.rect(0, 0, w, 0.8 * cm, fill=1, stroke=0)
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica", 7.5)
        canvas.drawString(1.6 * cm, 0.28 * cm,
                          "Seeding QDArchive (SQ26) · FAU Erlangen-Nürnberg")
        canvas.drawRightString(w - 1.6 * cm, 0.28 * cm, f"Page {doc.page - 1}")
        canvas.restoreState()

    total_all = sum(v["total"] for v in views)
    classified_all = sum(v["classified"] for v in views)
    classifiable_all = sum(
        v["type_counts"].get(t, 0) for v in views for t in config.CLASSIFIABLE_TYPES
    )
    # Per-file classification coverage (primary files of the selected repos).
    frow = db.conn.execute(
        "SELECT COUNT(*) AS total, "
        "       SUM(CASE WHEN f.primary_class IS NOT NULL THEN 1 ELSE 0 END) AS done "
        "FROM files f JOIN projects p ON p.id = f.project_id "
        "WHERE LOWER(COALESCE(f.file_category,'')) = 'primary' "
        f"  AND p.source_repository IN ({','.join('?' * len(repos))})",
        repos,
    ).fetchone()
    files_primary_total = frow["total"] or 0
    files_classified = frow["done"] or 0

    _logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "data", "fau_logo_white.png")

    class Cover(Flowable):
        """Title page on a 2 cm margin grid: gold brand strip, eyebrow row with
        the FAU logo, centred-lowered title block, and a report-details panel."""

        def wrap(self, *a):
            return A4

        def draw(self):
            w, h = A4
            c = self.canv
            m = 2 * cm
            ice = HexColor("#B8D4F0")     # muted light blue for secondary text

            # Details-card geometry, defined first: the dark lower zone rises
            # to the card's vertical midpoint so the card straddles the two
            # background surfaces half-and-half.
            box_h = 6.7 * cm
            box_top = h * 0.475
            box_y = box_top - box_h

            # Background: navy with a darker lower zone; gold strip along the
            # top edge (mirrors the interior pages' header band language).
            c.setFillColor(navy)
            c.rect(0, 0, w, h, fill=1, stroke=0)
            c.setFillColor(HexColor("#0D1F35"))
            c.rect(0, 0, w, box_y + box_h / 2, fill=1, stroke=0)
            c.setFillColor(gold)
            c.rect(0, h - 0.18 * cm, w, 0.18 * cm, fill=1, stroke=0)

            # Eyebrow row: series label left, white FAU logo right
            c.setFillColor(gold)
            c.setFont("Helvetica-Bold", 10)
            #c.drawString(m, h - 2.1 * cm, "SEEDING QDARCHIVE · SQ26", charSpace=1.6)
            if os.path.exists(_logo_path):
                from reportlab.lib.utils import ImageReader
                iw, ih = ImageReader(_logo_path).getSize()
                lw = 5.6 * cm
                lh = lw * ih / iw
                c.drawImage(_logo_path, w - m - lw, h - 2.1 * cm - (lh - 0.28 * cm) / 2,
                            width=lw, height=lh, mask="auto")
            else:
                c.setFillColor(ice)
                c.setFont("Helvetica", 10)
                c.drawRightString(w - m, h - 2.1 * cm, "FAU Erlangen-Nürnberg")

            # Title block (lowered toward the visual centre of the page)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 38)
            c.drawString(m, h * 0.660, "QDArchive")
            c.setFont("Helvetica-Bold", 19)
            c.drawString(m, h * 0.612, "Data Classification Report")
            c.setStrokeColor(gold)
            c.setLineWidth(3)
            c.line(m, h * 0.588, w - m, h * 0.588)
            c.setFillColor(ice)
            c.setFont("Helvetica", 12)
            c.drawString(m, h * 0.552,
                         "ISIC Rev. 5 classification of qualitative research projects")

            # Report-details card — label-over-value grid (spec-sheet style),
            # half on the navy surface and half on the dark zone.
            c.setFillColor(HexColor("#0D1F35"))
            c.roundRect(m, box_y, w - 2 * m, box_h, 10, fill=1, stroke=0)
            c.setStrokeColor(gold)
            c.setLineWidth(1.4)
            c.roundRect(m, box_y, w - 2 * m, box_h, 10, fill=0, stroke=1)

            def field(x, y, label, value):
                c.setFillColor(gold)
                c.setFont("Helvetica-Bold", 8)
                c.drawString(x, y, label, charSpace=1.2)
                c.setFillColor(colors.white)
                c.setFont("Helvetica", 12.5)
                c.drawString(x, y - 0.55 * cm, value)

            col1_x, col2_x = 2.8 * cm, 10.9 * cm
            y = box_top - 1.0 * cm
            field(col1_x, y, "STUDENT NAME", "Mohammad Annus")
            field(col2_x, y, "STUDENT ID", "23221189")
            y -= 1.5 * cm
            field(col1_x, y, "SUPERVISOR", "Prof. Dr. Dirk Riehle")
            field(col2_x, y, "STANDARD", "ISIC Rev. 5 · division level")
            y -= 1.5 * cm
            field(col1_x, y, "DEPARTMENT",
                  "Open-Source Software · Department of Computer Science")
            y -= 1.5 * cm
            field(col1_x, y, "REPOSITORIES",
                  "Harvard Dataverse (#10) · Columbia Oral History Archive (#19)")

    # -- Document assembly ---------------------------------------------------
    w, h = A4
    margin = 1.8 * cm
    doc = BaseDocTemplate(path, pagesize=A4, leftMargin=margin, rightMargin=margin,
                          topMargin=1.8 * cm, bottomMargin=1.4 * cm)
    cover_frame = Frame(0, 0, w, h, leftPadding=0, rightPadding=0,
                        topPadding=0, bottomPadding=0)
    normal_frame = Frame(margin, 1.2 * cm, w - 2 * margin, h - 3.2 * cm,
                         leftPadding=0, rightPadding=0, topPadding=0.3 * cm, bottomPadding=0)
    doc.addPageTemplates([
        PageTemplate(id="cover", frames=[cover_frame]),
        PageTemplate(id="normal", frames=[normal_frame], onPage=on_page),
    ])

    story: list = [Cover(), NextPageTemplate("normal"), PageBreak()]

    # Executive summary
    story.append(Paragraph("Executive Summary", styles["h1"]))
    story.append(HRFlowable(width="100%", thickness=2, color=blue, spaceAfter=14))
    proj_pct = (classified_all / classifiable_all * 100) if classifiable_all else 0
    file_pct = (files_classified / files_primary_total * 100) if files_primary_total else 0
    story.append(Paragraph(
        f"This report presents the classification of <b>{total_all:,} qualitative "
        f"research projects</b> harvested from the two assigned repositories, "
        f"<b>Harvard Dataverse</b> (#10) and <b>Columbia Oral History Archive</b> "
        f"(#19). Each project was assigned a <b>project type</b> from its file "
        f"manifest and, where applicable, an <b>ISIC Rev. 5 division</b> (and its "
        f"primary data files individually) using a TF-IDF + cosine-similarity "
        f"classifier over the official ISIC Rev. 5 division corpus. Of the "
        f"<b>{classifiable_all:,}</b> classifiable projects (QDA + QD), "
        f"<b>{classified_all:,}</b> ({proj_pct:.1f}%) received a division, and "
        f"<b>{files_classified:,}</b> of {files_primary_total:,} primary data files "
        f"({file_pct:.1f}%) were classified individually; the rest were left "
        f"unclassified rather than forced into a weak match.", styles["body"]))
    story.append(Spacer(1, 0.15 * cm))

    # Headline analytics (all repositories combined)
    story.append(card_row([
        ("Projects harvested", f"{total_all:,}"),
        ("Classifiable (QDA + QD)", f"{classifiable_all:,}"),
        ("ISIC-classified projects", f"{classified_all:,}"),
        ("Primary files classified", f"{files_classified:,}"),
    ]))
    story.append(Spacer(1, 0.6 * cm))

    story.append(Paragraph("Project-type distribution (all repositories)", styles["h2"]))
    combined_types: dict[str, int] = {}
    for v in views:
        for t, c in v["type_counts"].items():
            combined_types[t] = combined_types.get(t, 0) + c
    crit = {
        config.PROJECT_TYPE_QDA: "Contains a QDA analysis file",
        config.PROJECT_TYPE_QD: "No QDA file but has primary data files",
        config.PROJECT_TYPE_OTHER: "Has files but no primary data files",
        config.PROJECT_TYPE_NONE: "No usable files identified",
    }
    trows = [["Project type", "Criterion", "Count"]]
    for t in config.PROJECT_TYPES:
        trows.append([t, crit[t], f"{combined_types.get(t, 0):,}"])
    tt = Table(trows, colWidths=[4.2 * cm, 9.3 * cm, 2.5 * cm])
    tt.setStyle(table_style([
        ("ALIGN", (2, 0), (2, -1), "CENTER"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 1), (0, -1), blue),
    ]))
    story.append(tt)
    story.append(Spacer(1, 0.6 * cm))

    # Composition per repository (100%-stacked bars; table above is the data view)
    story.append(Paragraph("Project-type composition by repository", styles["h2"]))
    comp = type_composition_chart(views)
    if comp:
        story.append(comp)
        story.append(Paragraph(
            "Figure: project-type share per repository (bar = 100% of that "
            "repository's projects; the number at the bar end is the project count).",
            styles["cap"]))
    story.append(Spacer(1, 0.5 * cm))

    # Classification coverage (single-ratio meters)
    story.append(Paragraph("Classification coverage", styles["h2"]))
    story.append(coverage_meters([
        ("Projects (of classifiable QDA + QD)", classified_all, classifiable_all),
        ("Primary data files", files_classified, files_primary_total),
    ]))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(
        "Unclassified remainder reflects the no-default-bucket policy: items whose "
        "metadata or file name carries too little signal stay NULL instead of being "
        "forced into a weak division.", styles["cap"]))
    story.append(PageBreak())

    # Per-repository sections
    for v in views:
        story.append(Paragraph(f"Repository: {v['name']}", styles["h1"]))
        story.append(HRFlowable(width="100%", thickness=2, color=blue, spaceAfter=14))
        story.append(stat_cards(v))
        story.append(Spacer(1, 0.45 * cm))

        # Project-type table
        story.append(Paragraph("Project-type distribution", styles["h2"]))
        rows = [["Project type", "Count", "Percentage"]]
        for t in config.PROJECT_TYPES:
            cnt = v["type_counts"].get(t, 0)
            if cnt:
                pct = (cnt / v["total"] * 100) if v["total"] else 0
                rows.append([t, f"{cnt:,}", f"{pct:.1f}%"])
        pt = Table(rows, colWidths=[7 * cm, 4.5 * cm, 4.5 * cm])
        pt.setStyle(table_style([
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, 1), (0, -1), blue),
        ]))
        story.append(pt)
        story.append(Spacer(1, 0.4 * cm))

        # Histogram (vector) — the heading carries the title; the chart itself is untitled
        story.append(Paragraph(
            f"Primary ISIC class distribution : {v['name']} (top {top_n})",
            styles["h2"]))
        chart = hbar_chart(v["dist"])
        if chart:
            story.append(chart)
            story.append(Paragraph(
                f"Figure: top {min(top_n, v['n_classes'])} of {v['n_classes']} ISIC "
                f"divisions identified; the number at each bar end is the project count.",
                styles["cap"]))
        else:
            story.append(Paragraph("No ISIC-classified projects for this repository.",
                                   styles["body"]))
        story.append(Spacer(1, 0.3 * cm))

        # Top-20 ranked table — Paragraph cells wrap so the full class name
        # is always shown (spec 4d), never truncated.
        cell_style = ParagraphStyle("rankcell", fontName="Helvetica",
                                    fontSize=9, leading=11)
        rank_heading = Paragraph("Rank-ordered classes (top 20)", styles["h2"])
        rank_rows = [["Rank", "ISIC", "Division", "Count"]]
        for i, (code, cnt) in enumerate(v["dist"][:20], 1):
            name = name_map.get(code, code)
            rank_rows.append([str(i), code, Paragraph(name, cell_style), f"{cnt:,}"])
        if len(rank_rows) == 1:
            rank_rows.append(["—", "—", "No classes identified", "0"])
        rank_extra = [
            ("ALIGN", (0, 0), (1, -1), "CENTER"),
            ("ALIGN", (3, 0), (3, -1), "CENTER"),
            ("FONTNAME", (1, 1), (1, -1), "Helvetica-Bold"),
            ("TEXTCOLOR", (1, 1), (1, -1), blue),
        ]
        highlights = ["#FFF3CD", "#FFF8E6", "#FFFDF5"]
        for hi in range(1, min(4, len(rank_rows))):
            rank_extra.append(("BACKGROUND", (0, hi), (-1, hi), HexColor(highlights[hi - 1])))
        # repeatRows=1 re-prints the header row if the table is ever split;
        # KeepTogether stops the heading + first rows orphaning at a page bottom
        # (the whole block moves to the next page instead).
        rt = Table(rank_rows, colWidths=[1.5 * cm, 2 * cm, 10.5 * cm, 2 * cm],
                   repeatRows=1)
        rt.setStyle(table_style(rank_extra))
        story.append(KeepTogether([rank_heading, rt]))
        story.append(Spacer(1, 0.4 * cm))

        # Comments (kept as one block so the heading can't orphan)
        dom = display_class(v["dominant_code"], name_map) or "none"
        tc = v["type_counts"]
        story.append(KeepTogether([
            Paragraph("Comments on findings", styles["h2"]),
            Paragraph(
            f"The <b>{v['name']}</b> repository yielded "
            f"{tc.get(config.PROJECT_TYPE_QDA, 0)} <b>QDA_PROJECT</b> and "
            f"{tc.get(config.PROJECT_TYPE_QD, 0):,} <b>QD_PROJECT</b> entries, of which "
            f"{v['classified']:,} received an ISIC division across {v['n_classes']} "
            f"distinct classes. The dominant primary class is <b>{dom}</b> "
            f"({v['dominant_count']:,} projects). Classification uses metadata and file "
            f"names; because much of the metadata is sparse, some assignments are weak "
            f"lexical matches (each row carries a persisted confidence score), and "
            f"projects with no meaningful signal are intentionally left unclassified.",
            styles["body"]),
        ]))
        story.append(PageBreak())

    doc.build(story)
    logger.info("Wrote PDF → %s", path)
    return path


# ── Step 4b + Step 5: statistics + data-challenges summary ──────────────────
def print_statistics(db: Database, repos: list[str] | None = None) -> None:
    name_map = load_name_map()
    print("\n" + "=" * 60)
    print("  Classification statistics (per repository)")
    print("=" * 60)
    for repo_key in _repos(repos):
        cfg = config.REPOSITORIES.get(repo_key, {})
        stats = db.classification_stats(repo_key, config.CLASSIFIABLE_TYPES)
        print(f"\n[{cfg.get('name', repo_key)}]  (repository_id={cfg.get('repository_id')})")
        for t in (*config.PROJECT_TYPES, "UNCLASSIFIED"):
            if t in stats["type_counts"]:
                print(f"    {t:16s}: {stats['type_counts'][t]}")
        dom = stats["dominant_class"]
        if dom:
            print(f"    dominant class : {display_class(dom, name_map)} "
                  f"({stats['dominant_count']})")

    # Step 5 — technical challenges with data.
    challenges = db.get_challenges()
    print("\n" + "-" * 60)
    print(f"  Technical challenges (data) logged: {len(challenges)}")
    by_type: dict[str, int] = {}
    for c in challenges:
        key = c.get("challenge_type") or "other"
        by_type[key] = by_type.get(key, 0) + 1
    for ctype, n in sorted(by_type.items(), key=lambda kv: kv[1], reverse=True):
        print(f"    {ctype:28s}: {n}")
